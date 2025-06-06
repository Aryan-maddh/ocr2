# services/export_service.py
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
import uuid
import os
from typing import Dict, List, Optional

class ExportService:
    def __init__(self):
        self.export_dir = "exports"
        os.makedirs(self.export_dir, exist_ok=True)
    
    def export_document_data(
        self, 
        document_data: Dict, 
        format_type: str = "csv",
        include_metadata: bool = True,
        custom_template: Optional[Dict] = None
    ) -> str:
        """Export document data in various formats"""
        
        export_id = str(uuid.uuid4())
        
        if format_type.lower() == "csv":
            return self._export_to_csv(document_data, export_id, include_metadata)
        elif format_type.lower() == "xlsx":
            return self._export_to_excel(document_data, export_id, include_metadata)
        elif format_type.lower() == "pdf":
            return self._export_to_pdf(document_data, export_id, include_metadata)
        elif format_type.lower() == "json":
            return self._export_to_json(document_data, export_id, include_metadata)
        else:
            raise ValueError(f"Unsupported export format: {format_type}")
    
    def _export_to_csv(self, data: Dict, export_id: str, include_metadata: bool) -> str:
        """Export to CSV format"""
        rows = []
        
        if include_metadata:
            rows.append(["Document Type", data.get("document_type", "Unknown")])
            rows.append(["Processed At", data.get("processed_at", "")])
            rows.append(["Confidence Score", data.get("avg_confidence", "")])
            rows.append([])  # Empty row separator
        
        # Add field data
        rows.append(["Field Name", "Field Value", "Confidence"])
        
        for field_name, field_info in data.get("extracted_fields", {}).items():
            if isinstance(field_info, dict):
                rows.append([
                    field_name,
                    field_info.get("value", ""),
                    field_info.get("confidence", "")
                ])
            else:
                rows.append([field_name, str(field_info), ""])
        
        # Create DataFrame and save
        df = pd.DataFrame(rows)
        filename = f"{export_id}_document_data.csv"
        filepath = os.path.join(self.export_dir, filename)
        df.to_csv(filepath, index=False, header=False)
        
        return filepath
    
    def _export_to_excel(self, data: Dict, export_id: str, include_metadata: bool) -> str:
        """Export to Excel with formatting"""
        filename = f"{export_id}_document_data.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Document Data"
        
        # Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        row = 1
        
        if include_metadata:
            # Metadata section
            ws.merge_cells(f'A{row}:C{row}')
            ws[f'A{row}'] = "Document Information"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 1
            
            metadata_items = [
                ("Document Type", data.get("document_type", "Unknown")),
                ("Processed At", data.get("processed_at", "")),
                ("Average Confidence", data.get("avg_confidence", "")),
                ("Total Fields", len(data.get("extracted_fields", {})))
            ]
            
            for key, value in metadata_items:
                ws[f'A{row}'] = key
                ws[f'B{row}'] = str(value)
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1  # Empty row
        
        # Field data header
        headers = ["Field Name", "Field Value", "Confidence Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        row += 1
        
        # Field data
        for field_name, field_info in data.get("extracted_fields", {}).items():
            if isinstance(field_info, dict):
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = str(field_info.get("value", ""))
                ws[f'C{row}'] = field_info.get("confidence", "")
            else:
                ws[f'A{row}'] = field_name
                ws[f'B{row}'] = str(field_info)
                ws[f'C{row}'] = ""
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
        return filepath
    
    def _export_to_pdf(self, data: Dict, export_id: str, include_metadata: bool) -> str:
        """Export to PDF format"""
        filename = f"{export_id}_document_data.pdf"
        filepath = os.path.join(self.export_dir, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph("Document Data Export", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        if include_metadata:
            # Metadata section
            metadata_title = Paragraph("Document Information", styles['Heading2'])
            story.append(metadata_title)
            
            metadata_data = [
                ["Document Type", data.get("document_type", "Unknown")],
                ["Processed At", data.get("processed_at", "")],
                ["Average Confidence", data.get("avg_confidence", "")],
                ["Total Fields", str(len(data.get("extracted_fields", {})))]
            ]
            
            metadata_table = Table(metadata_data, colWidths=[2*72, 4*72])
            metadata_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(metadata_table)
            story.append(Spacer(1, 20))
        
        # Field data section
        fields_title = Paragraph("Extracted Fields", styles['Heading2'])
        story.append(fields_title)
        
        field_data = [["Field Name", "Field Value", "Confidence"]]
        
        for field_name, field_info in data.get("extracted_fields", {}).items():
            if isinstance(field_info, dict):
                field_data.append([
                    field_name,
                    str(field_info.get("value", ""))[:50] + "..." if len(str(field_info.get("value", ""))) > 50 else str(field_info.get("value", "")),
                    str(field_info.get("confidence", ""))
                ])
            else:
                field_data.append([
                    field_name,
                    str(field_info)[:50] + "..." if len(str(field_info)) > 50 else str(field_info),
                    ""
                ])
        
        field_table = Table(field_data, colWidths=[2*72, 3*72, 1*72])
        field_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        
        story.append(field_table)
        doc.build(story)
        
        return filepath
    
    def _export_to_json(self, data: Dict, export_id: str, include_metadata: bool) -> str:
        """Export to JSON format"""
        filename = f"{export_id}_document_data.json"
        filepath = os.path.join(self.export_dir, filename)
        
        export_data = {
            "export_id": export_id,
            "extracted_fields": data.get("extracted_fields", {}),
        }
        
        if include_metadata:
            export_data["metadata"] = {
                "document_type": data.get("document_type"),
                "processed_at": data.get("processed_at"),
                "avg_confidence": data.get("avg_confidence"),
                "total_fields": len(data.get("extracted_fields", {}))
            }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2, ensure_ascii=False)
        
        return filepath
    
    def batch_export_documents(
        self, 
        documents_data: List[Dict], 
        format_type: str = "xlsx",
        merge_into_single_file: bool = True
    ) -> str:
        """Export multiple documents into a single file or separate files"""
        
        if not merge_into_single_file:
            # Export each document separately
            export_paths = []
            for doc_data in documents_data:
                path = self.export_document_data(doc_data, format_type)
                export_paths.append(path)
            return export_paths
        
        # Merge into single file
        export_id = str(uuid.uuid4())
        
        if format_type.lower() == "xlsx":
            return self._batch_export_excel(documents_data, export_id)
        elif format_type.lower() == "csv":
            return self._batch_export_csv(documents_data, export_id)
        else:
            raise ValueError(f"Batch export not supported for format: {format_type}")
    
    def _batch_export_excel(self, documents_data: List[Dict], export_id: str) -> str:
        """Export multiple documents to single Excel file with multiple sheets"""
        filename = f"{export_id}_batch_export.xlsx"
        filepath = os.path.join(self.export_dir, filename)
        
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["Document", "Type", "Fields Count", "Avg Confidence", "Status"])
        
        for i, doc_data in enumerate(documents_data):
            doc_name = f"Document_{i+1}"
            
            # Add to summary
            summary_ws.append([
                doc_name,
                doc_data.get("document_type", "Unknown"),
                len(doc_data.get("extracted_fields", {})),
                doc_data.get("avg_confidence", ""),
                "Processed"
            ])
            
            # Create individual sheet for each document
            ws = wb.create_sheet(doc_name)
            
            # Headers
            ws.append(["Field Name", "Field Value", "Confidence"])
            
            # Data
            for field_name, field_info in doc_data.get("extracted_fields", {}).items():
                if isinstance(field_info, dict):
                    ws.append([
                        field_name,
                        str(field_info.get("value", "")),
                        field_info.get("confidence", "")
                    ])
                else:
                    ws.append([field_name, str(field_info), ""])
        
        wb.save(filepath)
        return filepath
    
    def _batch_export_csv(self, documents_data: List[Dict], export_id: str) -> str:
        """Export multiple documents to single CSV file"""
        filename = f"{export_id}_batch_export.csv"
        filepath = os.path.join(self.export_dir, filename)
        
        all_rows = []
        all_rows.append(["Document_ID", "Document_Type", "Field_Name", "Field_Value", "Confidence"])
        
        for i, doc_data in enumerate(documents_data):
            doc_id = f"Document_{i+1}"
            doc_type = doc_data.get("document_type", "Unknown")
            
            for field_name, field_info in doc_data.get("extracted_fields", {}).items():
                if isinstance(field_info, dict):
                    all_rows.append([
                        doc_id,
                        doc_type,
                        field_name,
                        str(field_info.get("value", "")),
                        field_info.get("confidence", "")
                    ])
                else:
                    all_rows.append([doc_id, doc_type, field_name, str(field_info), ""])
        
        df = pd.DataFrame(all_rows)
        df.to_csv(filepath, index=False, header=False)
        return filepath

# Enhanced API endpoints for export
@app.post("/export_document/{document_id}")
async def export_document(
    document_id: str,
    format_type: str = "csv",
    include_metadata: bool = True,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export single document data"""
    export_service = ExportService()
    
    # Get document and extracted data
    document = db.query(Document).filter(
        Document.id == document_id,
        Document.user_id == current_user.id
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    extracted_data = db.query(ExtractedData).filter(
        ExtractedData.document_id == document_id
    ).all()
    
    # Prepare data for export
    export_data = {
        "document_type": document.document_type,
        "processed_at": document.processed_at.isoformat(),
        "extracted_fields": {}
    }
    
    confidences = []
    for data in extracted_data:
        export_data["extracted_fields"][data.field_name] = {
            "value": data.field_value,
            "confidence": float(data.confidence_score) if data.confidence_score else 0.0
        }
        if data.confidence_score:
            confidences.append(float(data.confidence_score))
    
    export_data["avg_confidence"] = sum(confidences) / len(confidences) if confidences else 0.0
    
    # Export
    filepath = export_service.export_document_data(export_data, format_type, include_metadata)
    
    return FileResponse(
        filepath,
        media_type="application/octet-stream",
        filename=os.path.basename(filepath)
    )

@app.post("/batch_export/")
async def batch_export_documents(
    document_ids: List[str],
    format_type: str = "xlsx",
    merge_files: bool = True,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export multiple documents"""
    export_service = ExportService()
    
    documents_data = []
    
    for doc_id in document_ids:
        document = db.query(Document).filter(
            Document.id == doc_id,
            Document.user_id == current_user.id
        ).first()
        
        if not document:
            continue
        
        extracted_data = db.query(ExtractedData).filter(
            ExtractedData.document_id == doc_id
        ).all()
        
        export_data = {
            "document_id": doc_id,
            "document_type": document.document_type,
            "processed_at": document.processed_at.isoformat(),
            "extracted_fields": {}
        }
        
        for data in extracted_data:
            export_data["extracted_fields"][data.field_name] = {
                "value": data.field_value,
                "confidence": float(data.confidence_score) if data.confidence_score else 0.0
            }
        
        documents_data.append(export_data)
    
    if not documents_data:
        raise HTTPException(status_code=404, detail="No documents found")
    
    filepath = export_service.batch_export_documents(documents_data, format_type, merge_files)
    
    return FileResponse(
        filepath,
        media_type="application/octet-stream",
        filename=os.path.basename(filepath)
    )